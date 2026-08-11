"""

32-2. 對比excel與資料夾的差別(萬用版).py

"""

import os
from openpyxl import load_workbook


def s_folder():
    """找出空資料夾"""
    folder_path = input("放入資料夾(用拖拉路徑的方式) : ")
    folder_path = os.path.relpath(folder_path)
    lst = []
    for x, y, _ in os.walk(folder_path):
        if y == []:
            lst.append(x)
    return lst


def v_excel() -> None:
    """excel下的路徑"""
    result = []
    lst_excel = []

    wb = load_workbook(input("放入excel檔案(用拖拉路徑的方式即可) : "))
    excel_path = input("在拉一次資料夾(用拖拉路徑的方式) : ")
    excel_path = os.path.relpath(excel_path)

    sheet_name = input("輸入分頁名稱 : ")
    ws = wb[sheet_name]  # 也可以直接寫分頁名稱

    for row in ws.iter_rows():  # 從第一個row開始讀, 在讀第2、3...row
        # list comprehension
        result.append([cell.value for cell in row])

    for r in result[
        1:
    ]:  # 從第2個 (index = 1) row開始抓, 第1個row是每一個column 的標題所以不抓
        # total += int(r[1])  # 要抓第12個值 (index = 11)
        a = rf"{excel_path}\{r[0]}\{r[1]}\{r[2]}\{r[3]}"
        lst_excel.append(a)
    return lst_excel


# ########################################################

if __name__ == "__main__":
    try:
        set_s_floder = set(s_folder())
        set_v_excel = set(v_excel())

        s_v = set_s_floder - set_v_excel
        s_v = list(s_v)
        s_v.sort()
        print("######################")
        print("--------資料夾多了這些--------")
        print("######################")
        for i in s_v:
            print(i)
        print(f"{len(s_v)} 個")
        print("######################")

        v_s = set_v_excel - set_s_floder
        v_s = list(v_s)
        v_s.sort()
        print("--------excel多了這些--------")
        print("######################")
        for j in v_s:
            print(j)
        print(f"{len(v_s)} 個")
        print("######################")
        input("Press Enter to Exit")
    except Exception as e:
        print(e)
        input("Press Enter to Exit")
